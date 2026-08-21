# -*- coding: utf-8 -*-
"""
產生「連線異常查詢案件登記系統」Phase 1 的 Excel 範本與測試用範例資料。

用法（一般使用者不需要執行，檔案已經產生好放在上一層資料夾）：
    pip install openpyxl
    python 產生Excel範本與範例資料.py
"""
import os, random, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# ---------------------------------------------------------------- 欄位定義
CASE_COLUMNS = [
    ('案件編號',            12, None),
    ('登記時間',            18, 'yyyy/mm/dd hh:mm'),
    ('申請人／單位',        18, None),
    ('問題類型',            30, None),
    ('問題類型其他說明',    22, None),
    ('來源 IP（含網段）',   22, None),
    ('目的 IP 或網域',      24, None),
    ('目的通訊埠',          12, None),
    ('通訊協定',            12, None),
    ('錯誤訊息原文',        40, None),
    ('發生時間',            18, 'yyyy/mm/dd hh:mm'),
    ('用途說明',            26, None),
    ('預計使用期間',        14, None),
    ('實際原因',            34, None),
    ('實際原因其他說明',    22, None),
    ('責任單位',            30, None),
    ('責任單位其他說明',    22, None),
    ('處理耗時',            12, None),
    ('是否重複案件',        14, None),
    ('處理狀態',            12, None),
    ('結案時間',            18, 'yyyy/mm/dd hh:mm'),
    ('備註',                26, None),
]

OPTIONS = {
    '申請人／單位': ['核心系統科', '網銀開發科', '信用卡系統科', '外匯系統科', '數位金融科', '委外廠商'],
    '問題類型': [
        '連線逾時，完全不通',
        '連線被拒絕（Connection refused / Reset）',
        '網頁功能異常，疑似被擋（送出失敗、特定參數出錯）',
        '檔案上傳失敗',
        '對外網站或外部 API 無法存取',
        '憑證／加密相關錯誤（SSL/TLS）',
        '其他（需填說明）',
    ],
    '通訊協定': ['TCP', 'UDP', 'HTTP', 'HTTPS', '其他'],
    '預計使用期間': ['一次性', '長期'],
    '實際原因': [
        'WAF 規則誤判／攔截',
        'WebGateway 網址分類阻擋',
        'Arbor DDoS 防禦系統流量清洗',
        'FortiADC 設定（系統組）',
        'Palo Alto 防火牆政策未開通（系統組）',
        'Forti 系列防火牆政策未開通（301E／60F／61F，系統組）',
        'SRX110 防火牆（系統組）',
        '路由／交換器問題（Cisco Router、CoreSwitch，系統組）',
        '專線問題（集保／中央外匯／票交所／JCIC／財金／NCCC／FXML）',
        '應用程式本身問題，非網路設備所致',
        '設定錯誤（IP／Port／網域打錯）',
        '其他（需填說明）',
    ],
    '責任單位': [
        '資安組（WAF、Arbor DDoS、WebGateway）',
        '系統組（FortiADC、Palo Alto、Forti 系列、SRX110、Cisco Router、CoreSwitch、專線）',
        '開發單位自行處理',
        '跨組（需填說明）',
    ],
    '處理耗時': ['5 分鐘內', '15 分鐘', '30 分鐘', '1 小時', '1 小時以上'],
    '處理狀態': ['待處理', '處理中', '已結案'],
}
OPTION_ORDER = ['申請人／單位', '問題類型', '通訊協定', '預計使用期間',
                '實際原因', '責任單位', '處理耗時', '處理狀態']

# 下拉來源欄位（案件登記的欄位標題 -> 選項工作表的欄位標題, 是否允許自由輸入）
DROPDOWNS = [
    ('申請人／單位', '申請人／單位', True),
    ('問題類型',     '問題類型',     False),
    ('通訊協定',     '通訊協定',     True),
    ('預計使用期間', '預計使用期間', False),
    ('實際原因',     '實際原因',     False),
    ('責任單位',     '責任單位',     False),
    ('處理耗時',     '處理耗時',     False),
    ('處理狀態',     '處理狀態',     False),
]

HEAD_FILL_SEC = PatternFill('solid', fgColor='1C5CAB')   # 資安組填寫欄位
HEAD_FILL_DEV = PatternFill('solid', fgColor='2A78D6')   # 開發人員填寫欄位
HEAD_FILL_SYS = PatternFill('solid', fgColor='57554E')   # 系統自動
HEAD_FONT = Font(color='FFFFFF', bold=True, size=11)
THIN = Side(style='thin', color='D9D9D9')

DEV_FIELDS = {'申請人／單位', '問題類型', '問題類型其他說明', '來源 IP（含網段）', '目的 IP 或網域',
              '目的通訊埠', '通訊協定', '錯誤訊息原文', '發生時間', '用途說明', '預計使用期間'}
SEC_FIELDS = {'實際原因', '實際原因其他說明', '責任單位', '責任單位其他說明', '處理耗時',
              '是否重複案件', '處理狀態', '結案時間', '備註'}


def build_option_sheet(ws):
    ws.title = '選項'
    ws.freeze_panes = 'A2'
    for i, name in enumerate(OPTION_ORDER, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = PatternFill('solid', fgColor='1BAF7A')
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(46, len(name) * 2 + 12))
        for j, v in enumerate(OPTIONS[name], start=2):
            ws.cell(row=j, column=i, value=v).border = Border(bottom=THIN)
    ws.row_dimensions[1].height = 24
    note_row = max(len(v) for v in OPTIONS.values()) + 3
    ws.cell(row=note_row, column=1,
            value='※ 選項可自行往下新增，控制台會自動讀取；請勿更動第 1 列的欄位標題，也不要在中間留空白列。'
            ).font = Font(color='8B887F', size=10)


def build_case_sheet(ws):
    ws.title = '案件登記'
    ws.freeze_panes = 'A2'
    for i, (name, width, numfmt) in enumerate(CASE_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = HEAD_FILL_DEV if name in DEV_FIELDS else (HEAD_FILL_SEC if name in SEC_FIELDS else HEAD_FILL_SYS)
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = 'A1:%s1' % get_column_letter(len(CASE_COLUMNS))
    for i, (name, width, numfmt) in enumerate(CASE_COLUMNS, start=1):
        if numfmt:
            for r in range(2, 1002):
                ws.cell(row=r, column=i).number_format = numfmt


def add_validations(ws, last_row=1000):
    headers = [c[0] for c in CASE_COLUMNS]
    for case_col, opt_col, allow_free in DROPDOWNS:
        ci = headers.index(case_col) + 1
        oi = OPTION_ORDER.index(opt_col) + 1
        letter = get_column_letter(oi)
        formula = "='選項'!$%s$2:$%s$200" % (letter, letter)
        dv = DataValidation(type='list', formula1=formula, allow_blank=True,
                            showErrorMessage=not allow_free)
        dv.errorTitle = '請由下拉選單選擇'
        dv.error = '此欄僅能選擇「選項」工作表中列出的項目；若需要新的項目，請到「選項」工作表最下方新增。'
        dv.promptTitle = case_col
        dv.prompt = '請由下拉選單選擇' + ('（可自行輸入新的名稱）' if allow_free else '')
        ws.add_data_validation(dv)
        col = get_column_letter(ci)
        dv.add('%s2:%s%d' % (col, col, last_row))

    ci = headers.index('是否重複案件') + 1
    dv2 = DataValidation(type='list', formula1='"是,否"', allow_blank=True, showErrorMessage=True)
    dv2.promptTitle = '是否重複案件'
    dv2.prompt = '同一系統／同一人反覆詢問相同問題時選「是」。'
    ws.add_data_validation(dv2)
    col = get_column_letter(ci)
    dv2.add('%s2:%s%d' % (col, col, last_row))


def make_template(path):
    wb = Workbook()
    build_case_sheet(wb.active)
    build_option_sheet(wb.create_sheet())
    add_validations(wb['案件登記'])
    wb.save(path)
    print('已產生範本：', path)


# ---------------------------------------------------------------- 範例資料
def make_sample(xlsx_path, csv_path):
    random.seed(20260821)
    headers = [c[0] for c in CASE_COLUMNS]
    types = OPTIONS['問題類型']
    causes = OPTIONS['實際原因']
    owners = OPTIONS['責任單位']
    times = OPTIONS['處理耗時']
    apps = OPTIONS['申請人／單位']

    # 症狀 -> 常見原因（讓範例資料的對照表看起來合理）
    likely = {
        types[0]: [causes[4], causes[5], causes[7], causes[3]],
        types[1]: [causes[4], causes[10], causes[9], causes[6]],
        types[2]: [causes[0], causes[0], causes[9], causes[10]],
        types[3]: [causes[0], causes[3], causes[9]],
        types[4]: [causes[1], causes[1], causes[8], causes[4]],
        types[5]: [causes[9], causes[0], causes[10]],
        types[6]: [causes[11], causes[9]],
    }
    cause_owner = {
        causes[0]: owners[0], causes[1]: owners[0], causes[2]: owners[0],
        causes[3]: owners[1], causes[4]: owners[1], causes[5]: owners[1],
        causes[6]: owners[1], causes[7]: owners[1], causes[8]: owners[1],
        causes[9]: owners[2], causes[10]: owners[2], causes[11]: owners[3],
    }
    dsts = ['10.1.103.24', '10.1.101.15', 'api.twse.com.tw', '10.2.100.31', 'gw.nccc.com.tw',
            '192.168.241.10', 'ocsp.twca.com.tw', '10.1.113.52']
    errs = ['curl: (28) Connection timed out after 30001 milliseconds',
            'java.net.ConnectException: Connection refused (Connection refused)',
            'ERR_CONNECTION_RESET',
            'HTTP 403 Forbidden — Request blocked by security policy',
            'javax.net.ssl.SSLHandshakeException: PKIX path building failed',
            'HTTP 502 Bad Gateway',
            'Upload failed: connection reset by peer']
    purposes = ['介接證交所行情資料', '網銀改版測試', '信用卡授權介接測試', '外匯匯率檔下載',
                'API 串接第三方支付', '報表檔案上傳至 FTP', '批次對帳檔傳輸']

    today = dt.datetime.now().replace(second=0, microsecond=0)
    rows = []
    n = 0
    # 過去 5 個月的資料，件數逐月增加
    for months_ago in range(5, -1, -1):
        base = (today.replace(day=1) - dt.timedelta(days=1)).replace(day=1) if False else None
        y = today.year
        m = today.month - months_ago
        while m <= 0:
            m += 12
            y -= 1
        count = [6, 8, 7, 11, 12, 9][5 - months_ago]
        for _ in range(count):
            n += 1
            day = random.randint(1, 27)
            if months_ago == 0:
                day = min(day, max(1, today.day))
            reg = dt.datetime(y, m, day, random.randint(9, 17), random.choice([0, 5, 12, 20, 33, 41, 55]))
            if reg > today:
                reg = today - dt.timedelta(hours=random.randint(1, 40))
            t = random.choice(types[:6] + [types[random.randint(0, 6)]])
            cause = random.choice(likely.get(t, causes))
            owner = cause_owner[cause]
            # 最近幾天的案件保留未結案，作為待辦區示範
            open_case = (today - reg) < dt.timedelta(days=9) and random.random() < 0.8
            row = {
                '案件編號': 'C%s%03d' % (reg.strftime('%y%m'), n),
                '登記時間': reg,
                '申請人／單位': random.choice(apps),
                '問題類型': t,
                '問題類型其他說明': '批次排程無法連線' if t == types[6] else '',
                '來源 IP（含網段）': random.choice(['192.168.251.21（開發 Server）', '192.168.253.88（開發 PC）',
                                                   '192.168.241.7（開發 DMZ）', '10.1.113.30（測試環境）']),
                '目的 IP 或網域': random.choice(dsts),
                '目的通訊埠': random.choice([443, 80, 8080, 1521, 22, 3306, 21]),
                '通訊協定': random.choice(['TCP', 'HTTPS', 'HTTPS', 'HTTP', 'TCP']),
                '錯誤訊息原文': random.choice(errs),
                '發生時間': reg - dt.timedelta(minutes=random.randint(5, 90)),
                '用途說明': random.choice(purposes),
                '預計使用期間': random.choice(['一次性', '長期', '長期']),
                '實際原因': '' if open_case else cause,
                '實際原因其他說明': '',
                '責任單位': '' if open_case else owner,
                '責任單位其他說明': '',
                '處理耗時': '' if open_case else random.choice(times),
                '是否重複案件': '否' if open_case else random.choice(['否', '否', '否', '是']),
                '處理狀態': random.choice(['待處理', '處理中']) if open_case else '已結案',
                '結案時間': '' if open_case else (reg + dt.timedelta(hours=random.randint(1, 30))),
                '備註': '',
            }
            rows.append(row)
    rows.sort(key=lambda r: r['登記時間'])

    # 讓最近幾件維持未結案，示範「待辦區」的排序效果
    for k, row in enumerate(rows[-5:]):
        row['登記時間'] = today - dt.timedelta(days=6 - k, hours=random.randint(0, 6), minutes=random.randint(0, 55))
        row['發生時間'] = row['登記時間'] - dt.timedelta(minutes=random.randint(5, 60))
        row['實際原因'] = ''
        row['責任單位'] = ''
        row['處理耗時'] = ''
        row['結案時間'] = ''
        row['是否重複案件'] = '否'
        row['處理狀態'] = '處理中' if k % 3 == 0 else '待處理'
    rows.sort(key=lambda r: r['登記時間'])

    wb = Workbook()
    ws = wb.active
    build_case_sheet(ws)
    build_option_sheet(wb.create_sheet())
    add_validations(ws)
    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(headers, start=1):
            v = row.get(h, '')
            if v != '':
                ws.cell(row=i, column=j, value=v)
        for j, (name, width, numfmt) in enumerate(CASE_COLUMNS, start=1):
            if numfmt:
                ws.cell(row=i, column=j).number_format = numfmt
    wb.save(xlsx_path)
    print('已產生範例資料：', xlsx_path, '共', len(rows), '筆')

    import csv
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(headers)
        for row in rows:
            w.writerow([row[h].strftime('%Y/%m/%d %H:%M') if isinstance(row[h], dt.datetime) else row[h]
                        for h in headers])
    print('已產生範例資料：', csv_path)


if __name__ == '__main__':
    make_template(os.path.join(OUT, '案件登記_範本.xlsx'))
    make_sample(os.path.join(OUT, '範例資料_測試用.xlsx'),
                os.path.join(OUT, '範例資料_測試用.csv'))
